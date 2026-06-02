"""Service layer for Exam Planner domain logic."""

import csv
import io
import json
import random
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import Iterable, Optional
from uuid import UUID

from sqlalchemy import and_, func, or_, select
import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import (
    Chapter,
    ExamPaper,
    ExamPaperQuestion,
    Flashcard,
    GoalStatus,
    MistakeNotebook,
    QuestionBank,
    QuestionDifficulty,
    QuestionType,
    QuizAnswer,
    QuizAttempt,
    QuizMode,
    StudyGoal,
    StudySession,
    Subject,
    UserStats,
)


def encode_options(options: list[str] | None) -> str | None:
    cleaned = [str(item).strip() for item in (options or []) if str(item).strip()]
    return json.dumps(cleaned, ensure_ascii=False) if cleaned else None


def decode_options(raw: str | None) -> list[str]:
    if not raw:
        return []
    try:
        value = json.loads(raw)
    except json.JSONDecodeError:
        return [item.strip() for item in raw.split("|") if item.strip()]
    return value if isinstance(value, list) else []


async def get_subject_or_none(db: AsyncSession, user_id: UUID, subject_id: int) -> Subject | None:
    result = await db.execute(select(Subject).where(Subject.id == subject_id, Subject.user_id == user_id))
    return result.scalars().first()


async def get_chapter_or_none(db: AsyncSession, user_id: UUID, chapter_id: int) -> Chapter | None:
    result = await db.execute(
        select(Chapter)
        .join(Subject, Subject.id == Chapter.subject_id)
        .where(Chapter.id == chapter_id, Subject.user_id == user_id)
    )
    return result.scalars().first()


async def assert_subject_chapter(
    db: AsyncSession,
    user_id: UUID,
    subject_id: int,
    chapter_id: Optional[int] = None,
) -> tuple[Subject, Chapter | None]:
    subject = await get_subject_or_none(db, user_id, subject_id)
    if not subject:
        raise ValueError("SUBJECT_NOT_FOUND")
    chapter = None
    if chapter_id is not None:
        chapter = await get_chapter_or_none(db, user_id, chapter_id)
        if not chapter or chapter.subject_id != subject_id:
            raise ValueError("CHAPTER_NOT_FOUND")
    return subject, chapter


async def list_subjects_with_counts(db: AsyncSession, user_id: UUID) -> list[dict]:
    subjects = list((await db.execute(select(Subject).where(Subject.user_id == user_id).order_by(Subject.name))).scalars().all())
    ids = [subject.id for subject in subjects]
    counts = {subject.id: {"chapter_count": 0, "flashcard_count": 0, "question_count": 0} for subject in subjects}
    if ids:
        for sid, count in (await db.execute(select(Chapter.subject_id, func.count(Chapter.id)).where(Chapter.subject_id.in_(ids)).group_by(Chapter.subject_id))).all():
            counts[sid]["chapter_count"] = count
        for sid, count in (await db.execute(select(Flashcard.subject_id, func.count(Flashcard.id)).where(Flashcard.subject_id.in_(ids)).group_by(Flashcard.subject_id))).all():
            counts[sid]["flashcard_count"] = count
        for sid, count in (await db.execute(select(QuestionBank.subject_id, func.count(QuestionBank.id)).where(QuestionBank.subject_id.in_(ids)).group_by(QuestionBank.subject_id))).all():
            counts[sid]["question_count"] = count
    return [{**subject.__dict__, **counts[subject.id]} for subject in subjects]


async def chapters_with_counts(db: AsyncSession, user_id: UUID, subject_id: int) -> list[dict]:
    await assert_subject_chapter(db, user_id, subject_id)
    chapters = list((await db.execute(select(Chapter).where(Chapter.subject_id == subject_id).order_by(Chapter.order_index, Chapter.id))).scalars().all())
    ids = [chapter.id for chapter in chapters]
    counts = {chapter.id: {"flashcard_count": 0, "question_count": 0} for chapter in chapters}
    if ids:
        for cid, count in (await db.execute(select(Flashcard.chapter_id, func.count(Flashcard.id)).where(Flashcard.chapter_id.in_(ids)).group_by(Flashcard.chapter_id))).all():
            counts[cid]["flashcard_count"] = count
        for cid, count in (await db.execute(select(QuestionBank.chapter_id, func.count(QuestionBank.id)).where(QuestionBank.chapter_id.in_(ids)).group_by(QuestionBank.chapter_id))).all():
            counts[cid]["question_count"] = count
    return [{**chapter.__dict__, **counts[chapter.id]} for chapter in chapters]


def question_to_response(question: QuestionBank) -> dict:
    return {
        "id": question.id,
        "user_id": question.user_id,
        "subject_id": question.subject_id,
        "chapter_id": question.chapter_id,
        "question_text": question.question_text,
        "question_type": question.question_type,
        "options": decode_options(question.options),
        "correct_answer": question.correct_answer,
        "explanation": question.explanation,
        "difficulty": question.difficulty,
        "created_at": question.created_at,
    }


async def parse_upload(filename: str, raw: bytes) -> list[dict]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if lower.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX_SUPPORT_NOT_INSTALLED") from exc
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]
    raise RuntimeError("INVALID_FILE_TYPE")


def _norm_row(row: dict) -> dict:
    return {str(key).strip().lower(): ("" if value is None else str(value).strip()) for key, value in row.items() if key is not None}


async def validate_import_rows(db: AsyncSession, user_id: UUID, import_type: str, rows: list[dict]) -> list[dict]:
    preview = []
    for index, raw in enumerate(rows, start=2):
        row = _norm_row(raw)
        errors = []
        subject_id = int(row["subject_id"]) if row.get("subject_id", "").isdigit() else None
        chapter_id = int(row["chapter_id"]) if row.get("chapter_id", "").isdigit() else None
        if not subject_id:
            errors.append({"row": index, "field": "subject_id", "message": "subject_id is required."})
        else:
            try:
                await assert_subject_chapter(db, user_id, subject_id, chapter_id)
            except ValueError:
                errors.append({"row": index, "field": "subject_id", "message": "Subject or chapter does not exist for this user."})

        if import_type == "flashcards":
            if not row.get("front") and not row.get("front_text"):
                errors.append({"row": index, "field": "front", "message": "front is required."})
            if not row.get("back") and not row.get("back_text"):
                errors.append({"row": index, "field": "back", "message": "back is required."})
        else:
            if not row.get("question_text"):
                errors.append({"row": index, "field": "question_text", "message": "question_text is required."})
            if not row.get("correct_answer"):
                errors.append({"row": index, "field": "correct_answer", "message": "correct_answer is required."})
            options = [item.strip() for item in row.get("options", "").split("|") if item.strip()]
            qtype = row.get("question_type", "MULTIPLE_CHOICE").upper()
            if qtype == "MULTIPLE_CHOICE" and len(options) < 2:
                errors.append({"row": index, "field": "options", "message": "Multiple choice questions need options separated by |."})
            if options and row.get("correct_answer") not in options:
                errors.append({"row": index, "field": "correct_answer", "message": "correct_answer must match one option."})
        preview.append({"row": index, "data": row, "errors": errors})
    return preview


async def commit_import(db: AsyncSession, user_id: UUID, import_type: str, rows: list[dict]) -> tuple[int, list[dict]]:
    preview = await validate_import_rows(db, user_id, import_type, rows)
    errors = [error for item in preview for error in item["errors"]]
    valid = [item["data"] for item in preview if not item["errors"]]
    now = datetime.now(timezone.utc)
    if import_type == "flashcards":
        db.add_all([
            Flashcard(
                user_id=user_id,
                subject_id=int(row["subject_id"]),
                chapter_id=int(row["chapter_id"]) if row.get("chapter_id", "").isdigit() else None,
                front_text=row.get("front") or row.get("front_text"),
                back_text=row.get("back") or row.get("back_text"),
                tag=row.get("tag") or None,
                next_review_date=now,
            )
            for row in valid
        ])
    else:
        db.add_all([
            QuestionBank(
                user_id=user_id,
                subject_id=int(row["subject_id"]),
                chapter_id=int(row["chapter_id"]) if row.get("chapter_id", "").isdigit() else None,
                question_text=row["question_text"],
                question_type=QuestionType(row.get("question_type", "MULTIPLE_CHOICE").upper()),
                options=encode_options([item.strip() for item in row.get("options", "").split("|") if item.strip()]),
                correct_answer=row["correct_answer"],
                explanation=row.get("explanation") or None,
                difficulty=QuestionDifficulty(row.get("difficulty", "MEDIUM").upper()),
            )
            for row in valid
        ])
    if valid:
        await db.commit()
    return len(valid), errors


async def list_questions(db: AsyncSession, user_id: UUID, subject_id: int | None, chapter_id: int | None, difficulty: QuestionDifficulty | None) -> list[QuestionBank]:
    query = select(QuestionBank).where(QuestionBank.user_id == user_id)
    if subject_id:
        query = query.where(QuestionBank.subject_id == subject_id)
    if chapter_id:
        query = query.where(QuestionBank.chapter_id == chapter_id)
    if difficulty:
        query = query.where(QuestionBank.difficulty == difficulty)
    return list((await db.execute(query.order_by(QuestionBank.created_at.desc()))).scalars().all())


async def build_question_quiz(db: AsyncSession, user_id: UUID, subject_id: int | None, chapter_id: int | None, difficulty: QuestionDifficulty | None, limit: int, exam_paper_id: int | None = None) -> list[QuestionBank]:
    if exam_paper_id:
        query = (
            select(QuestionBank)
            .join(ExamPaperQuestion, ExamPaperQuestion.question_id == QuestionBank.id)
            .join(ExamPaper, ExamPaper.id == ExamPaperQuestion.exam_paper_id)
            .where(ExamPaper.id == exam_paper_id, ExamPaper.user_id == user_id)
            .order_by(ExamPaperQuestion.order_index)
        )
        return list((await db.execute(query)).scalars().all())
    questions = await list_questions(db, user_id, subject_id, chapter_id, difficulty)
    random.shuffle(questions)
    return questions[:limit]


async def submit_quiz(db: AsyncSession, user_id: UUID, subject_id: int | None, chapter_id: int | None, mode: QuizMode, duration_seconds: int | None, answers: Iterable) -> tuple[QuizAttempt, list[dict]]:
    answers = list(answers)
    question_ids = [answer.question_id for answer in answers]
    questions = {
        question.id: question
        for question in (await db.execute(select(QuestionBank).where(QuestionBank.id.in_(question_ids), QuestionBank.user_id == user_id))).scalars().all()
    }
    attempt = QuizAttempt(user_id=user_id, subject_id=subject_id, chapter_id=chapter_id, mode=mode, total_questions=len(answers), duration_seconds=duration_seconds)
    db.add(attempt)
    await db.flush()
    results = []
    correct_count = 0
    now = datetime.now(timezone.utc)
    for answer in answers:
        question = questions.get(answer.question_id)
        if not question:
            continue
        selected = (answer.selected_answer or "").strip()
        is_correct = selected.lower() == question.correct_answer.strip().lower()
        correct_count += 1 if is_correct else 0
        db.add(QuizAnswer(attempt_id=attempt.id, user_id=user_id, question_id=question.id, selected_answer=selected, correct_answer=question.correct_answer, is_correct=is_correct))
        await update_mistake(db, user_id, question, selected, is_correct, now)
        results.append({"question_id": question.id, "selected_answer": selected, "correct_answer": question.correct_answer, "is_correct": is_correct, "explanation": question.explanation})
    attempt.correct_count = correct_count
    attempt.wrong_count = max(0, len(results) - correct_count)
    attempt.total_questions = len(results)
    attempt.score = round((correct_count / len(results)) * 100, 2) if results else 0
    await db.commit()
    await db.refresh(attempt)
    return attempt, results


async def update_mistake(db: AsyncSession, user_id: UUID, question: QuestionBank, selected: str, is_correct: bool, now: datetime) -> None:
    result = await db.execute(select(MistakeNotebook).where(MistakeNotebook.user_id == user_id, MistakeNotebook.question_id == question.id))
    mistake = result.scalars().first()
    if is_correct:
        if mistake and mistake.resolved_at is None:
            mistake.correct_streak += 1
            if mistake.correct_streak >= 3:
                mistake.resolved_at = now
        return
    if mistake:
        mistake.selected_answer = selected
        mistake.correct_answer = question.correct_answer
        mistake.mistake_count += 1
        mistake.correct_streak = 0
        mistake.last_mistake_at = now
        mistake.resolved_at = None
    else:
        db.add(MistakeNotebook(user_id=user_id, question_id=question.id, selected_answer=selected, correct_answer=question.correct_answer, last_mistake_at=now))


async def dashboard(db: AsyncSession, user_id: UUID) -> dict:
    now = datetime.now(timezone.utc)
    week_start = now - timedelta(days=7)
    due = await db.scalar(select(func.count(Flashcard.id)).where(Flashcard.user_id == user_id, Flashcard.next_review_date <= now)) or 0
    mistakes = await db.scalar(select(func.count(MistakeNotebook.id)).where(MistakeNotebook.user_id == user_id, MistakeNotebook.resolved_at.is_(None))) or 0
    study_week = await db.scalar(select(func.coalesce(func.sum(StudySession.minutes), 0)).where(StudySession.user_id == user_id, StudySession.studied_at >= week_start)) or 0
    stats = await db.get(UserStats, user_id)
    exams = [
        {"subject_id": s.id, "name": s.name, "exam_date": s.exam_date, "days_left": max(0, (s.exam_date.date() - now.date()).days)}
        for s in (await db.execute(select(Subject).where(Subject.user_id == user_id, Subject.exam_date.is_not(None)).order_by(Subject.exam_date).limit(5))).scalars().all()
    ]
    goals = [
        {"id": g.id, "title": g.title, "subject_id": g.subject_id, "progress": round((g.completed_minutes / g.target_minutes) * 100, 1) if g.target_minutes else 0, "deadline": g.deadline}
        for g in (await db.execute(select(StudyGoal).where(StudyGoal.user_id == user_id, StudyGoal.status == GoalStatus.ACTIVE).order_by(StudyGoal.deadline.asc().nullslast()).limit(5))).scalars().all()
    ]
    weak, subject_accuracy, recs = await accuracy_rollups(db, user_id)
    return {
        "due_flashcards": due,
        "exam_countdown": exams,
        "current_streak": stats.current_streak if stats else 0,
        "longest_streak": stats.longest_streak if stats else 0,
        "weak_chapters": weak,
        "mistake_count": mistakes,
        "study_time_this_week": study_week,
        "goal_progress": goals,
        "recommendations": recs,
        "subject_accuracy": subject_accuracy,
    }


async def accuracy_rollups(db: AsyncSession, user_id: UUID) -> tuple[list[dict], list[dict], list[str]]:
    query = (
        select(QuestionBank.subject_id, QuestionBank.chapter_id, func.count(QuizAnswer.id), func.sum(func.cast(QuizAnswer.is_correct, sa.Integer)))
        .join(QuestionBank, QuestionBank.id == QuizAnswer.question_id)
        .where(QuizAnswer.user_id == user_id)
        .group_by(QuestionBank.subject_id, QuestionBank.chapter_id)
    )
    rows = (await db.execute(query)).all()
    subject_totals = defaultdict(lambda: [0, 0])
    weak = []
    recs = []
    for sid, cid, total, correct in rows:
        correct = int(correct or 0)
        total = int(total or 0)
        subject_totals[sid][0] += correct
        subject_totals[sid][1] += total
        accuracy = round((correct / total) * 100, 1) if total else 0
        if cid and accuracy < 60:
            chapter = await db.get(Chapter, cid)
            weak.append({"subject_id": sid, "chapter_id": cid, "chapter_title": chapter.title if chapter else f"Chapter {cid}", "accuracy": accuracy})
            recs.append(f"Accuracy chapter {chapter.title if chapter else cid} < 60%; ôn flashcard và luyện lại câu sai chương này.")
    subject_accuracy = [{"subject_id": sid, "accuracy": round((correct / total) * 100, 1) if total else 0} for sid, (correct, total) in subject_totals.items()]
    return weak[:8], subject_accuracy, recs[:8]


async def record_study_session(db: AsyncSession, user_id: UUID, payload) -> StudySession:
    if payload.subject_id:
        await assert_subject_chapter(db, user_id, payload.subject_id, payload.chapter_id)
    session = StudySession(user_id=user_id, **payload.model_dump())
    db.add(session)
    stats = await db.get(UserStats, user_id)
    today = date.today()
    if not stats:
        stats = UserStats(user_id=user_id, current_streak=1, longest_streak=1, total_study_minutes=0, last_study_date=today)
        db.add(stats)
    elif stats.last_study_date != today:
        stats.current_streak = stats.current_streak + 1 if stats.last_study_date == today - timedelta(days=1) else 1
        stats.longest_streak = max(stats.longest_streak, stats.current_streak)
        stats.last_study_date = today
    stats.total_study_minutes += payload.minutes
    if payload.goal_id:
        goal = await db.get(StudyGoal, payload.goal_id)
        if goal and goal.user_id == user_id:
            goal.completed_minutes += payload.minutes
            if goal.completed_minutes >= goal.target_minutes:
                goal.status = GoalStatus.COMPLETED
    await db.commit()
    await db.refresh(session)
    return session


async def create_exam_paper(db: AsyncSession, user_id: UUID, payload) -> tuple[ExamPaper, list[QuestionBank]]:
    await assert_subject_chapter(db, user_id, payload.subject_id)
    query = select(QuestionBank).where(QuestionBank.user_id == user_id, QuestionBank.subject_id == payload.subject_id)
    if payload.chapter_ids:
        query = query.where(QuestionBank.chapter_id.in_(payload.chapter_ids))
    questions = list((await db.execute(query)).scalars().all())
    if payload.difficulty_mix:
        selected = []
        by_diff = defaultdict(list)
        for question in questions:
            by_diff[question.difficulty.value].append(question)
        for diff, count in payload.difficulty_mix.items():
            pool = by_diff[diff.upper()]
            random.shuffle(pool)
            selected.extend(pool[: int(count)])
        questions = selected
    random.shuffle(questions)
    selected = questions[: payload.question_count]
    paper = ExamPaper(user_id=user_id, subject_id=payload.subject_id, title=payload.title, duration_minutes=payload.duration_minutes, question_count=len(selected))
    db.add(paper)
    await db.flush()
    db.add_all([ExamPaperQuestion(exam_paper_id=paper.id, question_id=q.id, order_index=i) for i, q in enumerate(selected, start=1)])
    await db.commit()
    await db.refresh(paper)
    return paper, selected


async def exam_paper_questions(db: AsyncSession, user_id: UUID, paper_id: int) -> tuple[ExamPaper | None, list[QuestionBank]]:
    paper = (await db.execute(select(ExamPaper).where(ExamPaper.id == paper_id, ExamPaper.user_id == user_id))).scalars().first()
    if not paper:
        return None, []
    questions = list((await db.execute(select(QuestionBank).join(ExamPaperQuestion, ExamPaperQuestion.question_id == QuestionBank.id).where(ExamPaperQuestion.exam_paper_id == paper.id).order_by(ExamPaperQuestion.order_index))).scalars().all())
    return paper, questions
